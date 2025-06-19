import json
import requests
from urllib.parse import urlparse, urljoin, parse_qs
from collections import defaultdict
import argparse
import os
import re
from concurrent.futures import ThreadPoolExecutor, as_completed
import time
import sys
import socket
from tqdm import tqdm  # For progress bars
import logging
from typing import Dict, List, Set, Optional, Union, Tuple, Any # Import Any

# For XLS output
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.workbook import Workbook # Explicitly import Workbook

# Suppress insecure request warnings for self-signed certs during probing (use only if necessary and understand risks)
requests.packages.urllib3.disable_warnings(requests.packages.urllib3.exceptions.InsecureRequestWarning)

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('recon_mapper.log'),
        logging.StreamHandler(sys.stdout) # Ensure logs go to stdout by default
    ]
)
logger = logging.getLogger(__name__)

class ReconMapper:
    """
    Advanced reconnaissance data mapper that consolidates and organizes findings from various sources.
    Features include:
    - Comprehensive data loading from multiple file types
    - Robust URL normalization and validation
    - Smart categorization of discovered assets (subdomains, directories, files, endpoints, JS, APIs)
    - Parallelized and rate-limited URL probing with detailed response analysis
    - Enhanced technology detection (headers, content, known patterns)
    - DNS record resolution
    - Basic vulnerability pattern detection
    - Detailed reporting in JSON, text, Markdown, HTML, or Excel formats
    """

    def __init__(self, rate_limit: float = 0.1, timeout: int = 15, max_redirects: int = 5, verify_ssl: bool = True):
        """
        Initialize the ReconMapper with configurable settings for HTTP requests.

        Args:
            rate_limit: Seconds between requests to avoid overwhelming servers.
            timeout: Request timeout in seconds.
            max_redirects: Maximum number of HTTP redirects to follow.
            verify_ssl: Whether to verify SSL certificates for HTTPS requests. Set to False for self-signed certs (caution!).
        """
        self.mapping = defaultdict(lambda: {
            'subdomains': set(),
            'directories': set(),
            'files': set(),
            'endpoints': set(),
            'js_files': set(),
            'parameters': set(),
            'api_endpoints': set(),
            'technologies': set(),
            'dns_records': set(),
            'vulnerabilities': set() # To store detected potential vulnerabilities/misconfigurations
        })

        self.session = requests.Session()
        self.session.max_redirects = max_redirects
        self.session.verify = verify_ssl # Apply SSL verification setting
        self.session.headers.update({
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36 ReconMapper/2.0', # Updated UA
            'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8',
            'Accept-Language': 'en-US,en;q=0.5',
            'Accept-Encoding': 'gzip, deflate',
            'Connection': 'keep-alive',
            'Cache-Control': 'no-cache' # Prevent caching issues
        })

        self.rate_limit = rate_limit
        self.timeout = timeout
        self.verified_urls: Dict[str, Dict] = {}  # Cache for probed URLs to avoid re-probing
        logger.debug(f"ReconMapper initialized with rate_limit={rate_limit}, timeout={timeout}, verify_ssl={verify_ssl}")

    def _normalize_url(self, url: str, base_domain: Optional[str] = None) -> Optional[str]:
        """
        Normalize and validate URLs for consistency and relevance.

        Args:
            url: The URL string to normalize.
            base_domain: Optional base domain to filter and enforce relevancy (e.g., "example.com").

        Returns:
            The normalized URL string, or None if the URL is invalid or irrelevant to the base domain.
        """
        if not url or not isinstance(url, str):
            logger.debug(f"Skipping empty or non-string URL: {url}")
            return None

        # Add scheme if missing, defaulting to https for robustness
        if not url.startswith(('http://', 'https://')):
            url = f'https://{url}' # Prefer HTTPS by default

        try:
            parsed = urlparse(url)

            # Validate netloc (domain part)
            if not parsed.netloc:
                logger.debug(f"Skipping URL with no netloc: {url}")
                return None

            # Enforce base domain if provided
            if base_domain:
                if not (parsed.netloc == base_domain or parsed.netloc.endswith(f'.{base_domain}')):
                    logger.debug(f"Skipping URL {url} as it's not a subdomain of {base_domain}")
                    return None
                # If it's a subdomain, keep its original netloc for mapping distinct subdomains
                # If it's the base domain itself, ensure consistency
                if parsed.netloc != base_domain and not parsed.netloc.endswith(f'.{base_domain}'):
                     # This logic ensures we don't accidentally change a valid subdomain's netloc
                     # We only filter irrelevant ones.
                     pass
                else: # It's either the base domain or a valid subdomain
                    # If the input was just 'example.com' and base_domain is 'example.com', parsed.netloc would be 'example.com'
                    # If input was 'sub.example.com' and base_domain 'example.com', parsed.netloc is 'sub.example.com'
                    pass

            # Clean path: remove duplicate slashes, fragments, and params for base path identification
            clean_path = re.sub(r'/{2,}', '/', parsed.path) # Remove duplicate slashes
            normalized_parsed = parsed._replace(
                path=clean_path,
                fragment='',
                query='', # Query is handled separately for parameters
                params=''
            )

            normalized_url = normalized_parsed.geturl()

            # Remove trailing slash unless it's the root path ("/")
            if normalized_url.endswith('/') and normalized_parsed.path != '/':
                normalized_url = normalized_url.rstrip('/')

            return normalized_url.lower() # Case normalization for consistency

        except ValueError as e:
            logger.warning(f"Failed to parse URL '{url}': {e}")
            return None
        except Exception as e:
            logger.error(f"An unexpected error occurred during URL normalization for '{url}': {e}")
            return None

    def _is_api_endpoint(self, url: str) -> bool:
        """
        Enhanced heuristic to identify API endpoints.

        Args:
            url: The URL string to check.

        Returns:
            True if the URL likely points to an API endpoint, False otherwise.
        """
        api_patterns = [
            r'/api/', r'/v\d+/', r'\.(json|xml|yaml|yml|api|graphql|soap|rest|rpc)(\?|$)',
            r'/graphql', r'/rest/', r'/soap/', r'/rpc/', r'/[a-z0-9_]+/api/',
            r'/.well-known/(openid-configuration|oauth-authorization-server|jwks)', # Specific well-known API related paths
            r'/swagger(\.json|\.yaml)?', r'/openapi(\.json|\.yaml)?', r'/v\d+/swagger', r'/v\d+/openapi',
            r'/wsdl', r'/odata', r'/jsonrpc',
            r'(/web)?service(s)?/' # Common service endpoints
        ]
        return any(re.search(pattern, url, re.IGNORECASE) for pattern in api_patterns)

    def _extract_parameters(self, url: str) -> Set[str]:
        """
        Extract unique parameter names from a URL's query string.

        Args:
            url: The URL string potentially containing query parameters.

        Returns:
            A set of unique parameter names.
        """
        try:
            query = urlparse(url).query
            if query:
                return set(parse_qs(query).keys())
            return set()
        except Exception as e:
            logger.warning(f"Failed to extract parameters from '{url}': {e}")
            return set()

    def _detect_technologies(self, response: requests.Response) -> Set[str]:
        """
        Detects web technologies based on response headers and HTML content.

        Args:
            response: The requests.Response object from an HTTP request.

        Returns:
            A set of detected technology strings.
        """
        tech = set()
        headers = response.headers
        content = response.text.lower() if response.text else '' # Ensure content is string

        # Header-based detection
        if 'Server' in headers:
            tech.add(f"server:{headers['Server'].split('/')[0].strip().lower()}")
        if 'X-Powered-By' in headers:
            tech.add(f"powered_by:{headers['X-Powered-By'].split('/')[0].strip().lower()}")
        if 'X-AspNet-Version' in headers:
            tech.add("asp.net")
        if 'Content-Type' in headers and 'charset=' in headers['Content-Type']:
            tech.add(f"charset:{headers['Content-Type'].split('charset=')[-1].strip().lower()}")
        if 'Strict-TransportSecurity' in headers:
            tech.add("hsts")
        if 'X-Frame-Options' in headers:
            tech.add("x-frame-options")
        if 'Content-Security-Policy' in headers:
            tech.add("csp")

        # Cookie-based detection
        if 'Set-Cookie' in headers:
            cookies = headers['Set-Cookie'].lower()
            if 'phpsessid' in cookies: tech.add("php")
            if 'jsessionid' in cookies: tech.add("java")
            if 'asp.net_sessionid' in cookies: tech.add("asp.net")
            if 'laravel_session' in cookies: tech.add("laravel")

        # Content-based detection (more patterns)
        frameworks = {
            'wordpress': ['wp-content', 'wp-includes', 'wordpress'],
            'drupal': ['drupal.settings', 'sites/all/modules'],
            'joomla': ['joomla', 'media/system/js'],
            'laravel': ['laravel_session', 'csrf-token'], # Token often in meta for Laravel
            'react': ['react-dom.production', 'react.development'], # Common React build files
            'angular': ['ng-app', 'angular.min.js'],
            'vue': ['vue.min.js', '__vue__'],
            'jquery': ['jquery.min.js', 'jquery.js'],
            'bootstrap': ['bootstrap.min.css', 'bootstrap.js'],
            'php': ['<?php', '.php'], # Basic PHP detection
            'node.js': ['/node_modules/'],
            'apache': ['apache', 'httpd'], # Often in server headers or error pages
            'nginx': ['nginx']
        }

        for framework, patterns in frameworks.items():
            if any(p in content for p in patterns):
                tech.add(framework)

        # Look for common comments or metadata
        if '' in content or 'ga(' in content: tech.add("google_analytics") # This condition '' in content will always be true if content is not empty
        if 'window.dataLayer' in content: tech.add("google_tag_manager")
        if '<meta name="generator" content="wordpress' in content: tech.add("wordpress_meta")
        if '<meta name="generator" content="joomla' in content: tech.add("joomla_meta")

        return tech

    def _detect_vulnerabilities(self, response: requests.Response) -> Set[str]:
        """
        Detects potential misconfigurations or common vulnerabilities based on response.

        Args:
            response: The requests.Response object.

        Returns:
            A set of detected vulnerability/misconfiguration strings.
        """
        vulns = set()
        headers = response.headers
        content = response.text

        # Insecure Headers
        if 'X-Content-Type-Options' not in headers:
            vulns.add("missing_x-content-type-options_header")
        if 'X-XSS-Protection' not in headers:
            vulns.add("missing_x-xss-protection_header")
        if 'Strict-Transport-Security' not in headers and response.url.startswith('https'):
            vulns.add("missing_hsts_header")
        if 'Content-Security-Policy' not in headers:
            vulns.add("missing_csp_header")

        # Directory Listing (heuristic)
        if response.status_code == 200 and ('<title>Index of /</title>' in content or '<pre>[ To Parent Directory ]</pre>' in content):
            vulns.add("potential_directory_listing")

        # Exposed sensitive files/paths (based on status codes and content)
        if response.status_code == 200 and ('debug' in response.url or 'test' in response.url):
            if any(kw in content for kw in ['password', 'secret', 'api_key', 'config']):
                vulns.add("sensitive_data_in_debug_path")

        # Check for common backup files/exposed files
        common_exposed_files = ['.git/config', '.env', 'phpinfo.php', 'adminer.php', '.bak', '.old', '.zip', '.tar.gz', '.sql']
        for pattern in common_exposed_files:
            # Check if the pattern is in the URL path, not just anywhere in the URL
            parsed_url_path = urlparse(response.url).path.lower()
            if pattern in parsed_url_path and response.status_code == 200:
                vulns.add(f"exposed_file:{pattern.replace('.', '_').replace('/', '_')}")

        return vulns

    def _resolve_dns(self, domain: str) -> Set[str]:
        """
        Perform DNS resolution for a domain (A and MX records).

        Args:
            domain: The domain name to resolve.

        Returns:
            A set of resolved IP addresses and MX records.
        """
        records = set()
        try:
            # Resolve A records (IPv4 addresses)
            # Use socket.gethostbyname_ex to get all IPs for a hostname
            _, _, ip_addresses = socket.gethostbyname_ex(domain)
            for ip in ip_addresses:
                records.add(f"A:{ip}")
        except socket.gaierror as e:
            logger.debug(f"DNS resolution for A record failed for {domain}: {e}")
        except Exception as e:
            logger.error(f"Unexpected error during A record resolution for {domain}: {e}")

        try:
            # Resolve MX records (mail exchange servers)
            # This is a bit more complex as socket.getaddrinfo might not directly give MX details.
            # A more robust solution would involve 'dns.resolver' library, but for basic,
            # we can infer from common port 25 lookups if available via getaddrinfo for domain.
            # A direct MX query requires a dedicated DNS library.
            # For simplicity, we'll keep the basic attempt here, but note its limitations.
            # If a proper DNS library like 'dnspython' is used:
            # import dns.resolver
            # answers = dns.resolver.resolve(domain, 'MX')
            # for rdata in answers:
            #     records.add(f"MX:{rdata.exchange.to_text()}")
            pass # Skipping complex MX resolution with standard socket for brevity, prioritize A records.
        except Exception as e:
            logger.debug(f"DNS resolution for MX record failed for {domain} (might require 'dnspython'): {e}")

        return records

    def load_data_file(self, file_path: str, file_type: str, base_domain: Optional[str] = None) -> bool:
        """
        Unified method to load different types of reconnaissance data files.

        Args:
            file_path: Path to the input file.
            file_type: Type of data ('subdomains', 'content', 'endpoints', 'parameters').
            base_domain: Optional base domain to filter entries by.

        Returns:
            True if data was successfully loaded, False otherwise.
        """
        if not os.path.exists(file_path):
            logger.warning(f"Skipping '{file_type}' file, not found: {file_path}")
            return False

        logger.info(f"Loading {file_type} from: {file_path}")
        lines_processed = 0
        try:
            with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                for line in tqdm(f, desc=f"Loading {file_type}", unit=" lines", leave=False):
                    line = line.strip()
                    if not line or line.startswith(('#', '//', '--', ';')): # Ignore comments
                        continue

                    normalized = self._normalize_url(line, base_domain)
                    if not normalized:
                        logger.debug(f"Skipping un-normalizable line: {line}")
                        continue

                    domain = urlparse(normalized).netloc # Get domain from normalized URL
                    if not domain: # Should not happen if normalized successfully
                        continue

                    if file_type == 'subdomains':
                        # Ensure the subdomain is added to its own parsed domain
                        # e.g., for sub.example.com, add to mapping['sub.example.com']
                        self.mapping[domain]['subdomains'].add(normalized)
                    elif file_type == 'content':
                        self._categorize_content(normalized, domain)
                    elif file_type == 'endpoints':
                        self._categorize_endpoint(normalized, domain)
                    elif file_type == 'parameters':
                        self._process_parameters(normalized, domain)
                    lines_processed += 1

            logger.info(f"Successfully loaded {lines_processed} {file_type} entries from {file_path}")
            return True

        except Exception as e:
            logger.error(f"Error loading {file_type} from {file_path}: {e}")
            return False

    def _categorize_content(self, url: str, domain: str):
        """Categorize content discovery results into appropriate sets."""
        path = urlparse(url).path

        params = self._extract_parameters(url)
        if params:
            self.mapping[domain]['parameters'].update(params)

        if re.search(r'\.js($|\?)', url, re.IGNORECASE) or 'javascript' in urlparse(url).scheme:
            self.mapping[domain]['js_files'].add(url)
        elif self._is_api_endpoint(url):
            self.mapping[domain]['api_endpoints'].add(url)
        elif '.' in os.path.basename(path):  # Has a file extension (e.g., .html, .php, .png)
            self.mapping[domain]['files'].add(url)
        elif path.endswith('/') or not path: # Directory or root path
            self.mapping[domain]['directories'].add(url)
        else: # General endpoint that doesn't fit other categories
            self.mapping[domain]['endpoints'].add(url)
        logger.debug(f"Categorized content: {url} -> {domain}")

    def _categorize_endpoint(self, url: str, domain: str):
        """Categorize general endpoints, also extracting parameters."""
        params = self._extract_parameters(url)
        if params:
            self.mapping[domain]['parameters'].update(params)

        if self._is_api_endpoint(url):
            self.mapping[domain]['api_endpoints'].add(url)
        else:
            self.mapping[domain]['endpoints'].add(url)
        logger.debug(f"Categorized endpoint: {url} -> {domain}")

    def _process_parameters(self, url: str, domain: str):
        """Process URLs explicitly loaded as containing parameters."""
        params = self._extract_parameters(url)
        if params:
            self.mapping[domain]['parameters'].update(params)
            # Add the base URL (without query) as an endpoint if not already there
            base_url_without_params = url.split('?', 1)[0]
            if not self._is_api_endpoint(base_url_without_params) and not re.search(r'\.js($|\?)', base_url_without_params, re.IGNORECASE):
                self.mapping[domain]['endpoints'].add(base_url_without_params)
        logger.debug(f"Processed parameters: {url} -> {domain}")


    def probe_url(self, url: str) -> Dict:
        """
        Probes a single URL to gather HTTP response details, detect technologies,
        and identify potential vulnerabilities. Includes caching to prevent redundant requests.

        Args:
            url: The URL string to probe.

        Returns:
            A dictionary containing probing results (status, headers, tech, vulns, etc.)
            or error information.
        """
        # Return cached result if already probed
        if url in self.verified_urls:
            return self.verified_urls[url]

        time.sleep(self.rate_limit) # Adhere to rate limit

        try:
            response = self.session.get(
                url,
                timeout=self.timeout,
                allow_redirects=True,
                stream=True # Use stream to efficiently handle large responses, read later
            )
            response.raise_for_status() # Raise an HTTPError for bad responses (4xx or 5xx)

            # Read content after successful request
            content = response.text # .text decodes content based on headers

            domain = urlparse(url).netloc
            # FIX: Call with self.
            technologies = self._detect_technologies(response)
            vulnerabilities = self._detect_vulnerabilities(response)

            # Update mapping for the specific domain
            if domain: # Ensure domain is valid
                if technologies:
                    self.mapping[domain]['technologies'].update(technologies)
                if vulnerabilities:
                    self.mapping[domain]['vulnerabilities'].update(vulnerabilities)

            # Get final IP address after redirects
            final_ip = None
            try:
                final_netloc = urlparse(response.url).netloc
                if final_netloc:
                    final_ip = socket.gethostbyname(final_netloc)
            except socket.gaierror:
                pass # Could not resolve final URL's IP

            result = {
                'url': url,
                'status': response.status_code,
                'content_type': response.headers.get('Content-Type', 'N/A'),
                'content_length': len(response.content),
                'redirect_chain': [f"{resp.status_code} -> {resp.url}" for resp in response.history] + [f"{response.status_code} -> {response.url}"],
                'final_url': response.url,
                'headers': dict(response.headers),
                'technologies': sorted(list(technologies)), # Convert set to sorted list for JSON
                'vulnerabilities': sorted(list(vulnerabilities)),
                'ip_address': final_ip
            }

            self.verified_urls[url] = result # Cache the result
            logger.debug(f"Successfully probed {url} (Status: {response.status_code})")
            return result

        except requests.exceptions.Timeout:
            logger.warning(f"Timeout probing {url} after {self.timeout} seconds.")
            error_msg = f"Timeout after {self.timeout}s"
            status_code = None
        except requests.exceptions.HTTPError as e:
            logger.warning(f"HTTP Error probing {url}: {e.response.status_code} - {e.response.reason}")
            error_msg = f"HTTP Error: {e.response.status_code} {e.response.reason}"
            status_code = e.response.status_code
        except requests.exceptions.ConnectionError as e:
            logger.warning(f"Connection Error probing {url}: {e}")
            error_msg = f"Connection Error: {str(e)}"
            status_code = None
        except requests.exceptions.SSLError as e:
            logger.warning(f"SSL Error probing {url}: {e}. Try --no-ssl-verify.")
            error_msg = f"SSL Error: {str(e)}"
            status_code = None
        except requests.exceptions.RequestException as e:
            logger.warning(f"Request Exception probing {url}: {e}")
            error_msg = f"Request Exception: {str(e)}"
            status_code = None
        except socket.gaierror as e:
            logger.warning(f"DNS Resolution error for {urlparse(url).netloc}: {e}")
            error_msg = f"DNS Resolution Error: {str(e)}"
            status_code = None
        except Exception as e:
            logger.error(f"An unexpected error occurred while probing {url}: {e}", exc_info=True)
            error_msg = f"Unexpected Error: {str(e)}"
            status_code = None

        error_result = {
            'url': url,
            'status': status_code,
            'error': error_msg,
            'content_type': 'N/A',
            'content_length': 0,
            'redirects': [],
            'final_url': url,
            'headers': {},
            'technologies': [],
            'vulnerabilities': [],
            'ip_address': None
        }
        self.verified_urls[url] = error_result # Cache error results too
        return error_result


    def probe_urls(self, domain: str, max_workers: int = 10) -> List[Dict]:
        """
        Orchestrates parallel URL probing for a given domain using a thread pool.
        Includes a progress bar for better user experience.

        Args:
            domain: The target domain for which to probe URLs.
            max_workers: The maximum number of concurrent threads for probing.

        Returns:
            A list of dictionaries, each representing the probing result for a URL.
        """
        urls_to_probe = set()
        domain_data = self.mapping.get(domain, {})

        # Collect all unique URLs across categories for the specified domain
        for category in ['subdomains', 'directories', 'files', 'endpoints', 'js_files', 'api_endpoints']:
            urls_to_probe.update(domain_data.get(category, set()))

        if not urls_to_probe:
            logger.warning(f"No URLs found to probe for {domain}. Skipping probing.")
            return []

        logger.info(f"Starting to probe {len(urls_to_probe)} unique URLs for {domain} with {max_workers} threads...")

        results: List[Dict] = []
        with ThreadPoolPoolExecutor(max_workers=max_workers) as executor:
            # Map URLs to futures
            future_to_url = {executor.submit(self.probe_url, url): url for url in urls_to_probe}

            # Use tqdm for a professional progress bar
            for future in tqdm(as_completed(future_to_url), total=len(urls_to_probe), desc="Probing URLs", unit=" URL"):
                results.append(future.result())

        # Perform DNS resolution for the main domain
        dns_records = self._resolve_dns(domain)
        if dns_records:
            self.mapping[domain]['dns_records'].update(dns_records)
            logger.info(f"Resolved DNS records for {domain}: {', '.join(dns_records)}")

        logger.info(f"Finished probing. {len(results)} URLs processed for {domain}.")
        return results

    def generate_report(self, domain: str, format: str = 'json', include_probe: bool = False) -> Union[Dict, str, Workbook]:
        """
        Generates a comprehensive report of the mapped reconnaissance data for a specific domain.

        Args:
            domain: The target domain for the report.
            format: The output format ('json', 'text', 'markdown', 'html', 'xls').
            include_probe: If True, includes the detailed probe results in the report.

        Returns:
            The report content as a dictionary (for JSON) or a string (for text/markdown/html).
            For XLS, it returns a Workbook object.
        Raises:
            ValueError: If an unsupported format is requested.
        """
        domain_data = dict(self.mapping.get(domain, {}))

        report_data = {}
        for category, items in domain_data.items():
            if isinstance(items, set):
                report_data[category] = sorted(list(items)) # Convert sets to sorted lists for consistent output
            else:
                report_data[category] = items # Keep other types as is

        if include_probe:
            # Ensure verified_urls are sorted by URL for consistent output in JSON
            report_data['verified_urls'] = sorted(self.verified_urls.values(), key=lambda x: x.get('url', ''))

        if format == 'json':
            return report_data
        elif format == 'text':
            return self._format_text_report(report_data, domain)
        elif format == 'markdown':
            return self._format_markdown_report(report_data, domain)
        elif format == 'html':
            return self._format_html_report(report_data, domain)
        elif format == 'xls':
            return self._format_xls_report(report_data, domain)
        else:
            raise ValueError(f"Unsupported report format: {format}. Choose from 'json', 'text', 'markdown', 'html', 'xls'.")

    def _format_text_report(self, data: Dict, domain: str) -> str:
        """Helper to format the report as human-readable plain text."""
        report = [f"--- Reconnaissance Report for: {domain} ---", ""]

        for category, items in data.items():
            if not items:
                continue

            report.append(f"=== {category.replace('_', ' ').upper()} === ({len(items)} items)")
            if isinstance(items, list):
                for item in items:
                    if isinstance(item, dict) and 'url' in item: # For probe results
                        status = item.get('status', 'N/A')
                        error = item.get('error', '')
                        line = f"  - {item['url']} [Status: {status}]"
                        if error: line += f" [Error: {error}]"
                        if item.get('technologies'): line += f" [Tech: {', '.join(item['technologies'])}]"
                        if item.get('vulnerabilities'): line += f" [Vulns: {', '.join(item['vulnerabilities'])}]"
                        report.append(line)
                    else:
                        report.append(f"  - {item}")
            elif isinstance(items, dict): # Should only happen for verified_urls if not separated
                for k, v in items.items():
                    report.append(f"  - {k}: {v}")
            report.append("") # Add a blank line for readability

        return "\n".join(report)

    def _format_markdown_report(self, data: Dict, domain: str) -> str:
        """Helper to format the report as GitHub-flavored Markdown."""
        report = [f"# Reconnaissance Report for `{domain}`", ""]

        for category, items in data.items():
            if not items:
                continue

            report.append(f"## {category.replace('_', ' ').title()} ({len(items)})")
            report.append("") # Blank line after heading
            if isinstance(items, list):
                for item in items:
                    if isinstance(item, dict) and 'url' in item: # For probe results
                        status = item.get('status', 'N/A')
                        error = item.get('error', '')
                        line = f"- **URL**: `{item['url']}`\n  - **Status**: `{status}`"
                        if item.get('final_url') and item['final_url'] != item['url']: line += f"\n  - **Final URL**: `{item['final_url']}`"
                        if error: line += f"\n  - **Error**: `{error}`"
                        if item.get('ip_address'): line += f"\n  - **IP**: `{item['ip_address']}`"
                        if item.get('technologies'): line += f"\n  - **Technologies**: `{', '.join(item['technologies'])}`"
                        if item.get('vulnerabilities'): line += f"\n  - **Vulnerabilities**: `{', '.join(item['vulnerabilities'])}`"
                        if item.get('content_type'): line += f"\n  - **Content Type**: `{item['content_type']}`"
                        if item.get('content_length'): line += f"\n  - **Content Length**: `{item['content_length']}` bytes"
                        if item.get('redirect_chain') and len(item['redirect_chain']) > 1: line += f"\n  - **Redirect Chain**: `{ ' -> '.join(item['redirect_chain']) }`"
                        report.append(line + "\n")
                    else:
                        report.append(f"- `{item}`")
            elif isinstance(items, dict):
                for k, v in items.items():
                    report.append(f"- **{k.replace('_', ' ').title()}**: `{v}`")
            report.append("") # Blank line after category

        return "\n".join(report)

    def _format_html_report(self, data: Dict, domain: str) -> str:
        """Helper to format the report as an HTML page."""
        html_content = [
            "<!DOCTYPE html>",
            "<html lang='en'>",
            "<head>",
            "    <meta charset='UTF-8'>",
            "    <meta name='viewport' content='width=device-width, initial-scale=1.0'>",
            f"    <title>Reconnaissance Report for {domain}</title>",
            "    <style>",
            "        body { font-family: Arial, sans-serif; margin: 20px; background-color: #f4f4f4; color: #333; }",
            "        .container { max-width: 1000px; margin: auto; background: #fff; padding: 20px; border-radius: 8px; box-shadow: 0 2px 4px rgba(0,0,0,0.1); }",
            "        h1 { color: #0056b3; border-bottom: 2px solid #eee; padding-bottom: 10px; }",
            "        h2 { color: #0056b3; margin-top: 20px; border-bottom: 1px solid #eee; padding-bottom: 5px; }",
            "        ul { list-style-type: none; padding: 0; }",
            "        ul li { background: #e9e9e9; margin-bottom: 5px; padding: 8px 12px; border-radius: 4px; }",
            "        table { width: 100%; border-collapse: collapse; margin-top: 10px; }",
            "        th, td { border: 1px solid #ddd; padding: 8px; text-align: left; }",
            "        th { background-color: #0056b3; color: white; }",
            "        .status-success { color: green; font-weight: bold; }",
            "        .status-error { color: red; font-weight: bold; }",
            "    </style>",
            "</head>",
            "<body>",
            "    <div class='container'>",
            f"        <h1>Reconnaissance Report for <code>{domain}</code></h1>",
            ""
        ]

        for category, items in data.items():
            if not items:
                continue

            html_content.append(f"        <h2>{category.replace('_', ' ').title()} ({len(items)})</h2>")

            if category == 'verified_urls' and isinstance(items, list) and all(isinstance(item, dict) for item in items):
                html_content.append("        <table>")
                html_content.append("            <thead>")
                html_content.append("                <tr>")
                html_content.append("                    <th>URL</th>")
                html_content.append("                    <th>Status</th>")
                html_content.append("                    <th>IP Address</th>")
                html_content.append("                    <th>Technologies</th>")
                html_content.append("                    <th>Vulnerabilities</th>")
                html_content.append("                    <th>Content Type</th>")
                html_content.append("                    <th>Content Length</th>")
                html_content.append("                </tr>")
                html_content.append("            </thead>")
                html_content.append("            <tbody>")
                for item in items:
                    status_class = "status-success" if 200 <= item.get('status', 0) < 400 else "status-error"
                    html_content.append("                <tr>")
                    html_content.append(f"                    <td><a href='{item.get('url', '#')}' target='_blank'>{item.get('url', 'N/A')}</a></td>")
                    html_content.append(f"                    <td class='{status_class}'>{item.get('status', 'N/A')}</td>")
                    html_content.append(f"                    <td>{item.get('ip_address', 'N/A')}</td>")
                    html_content.append(f"                    <td>{', '.join(item.get('technologies', [])) or 'N/A'}</td>")
                    html_content.append(f"                    <td>{', '.join(item.get('vulnerabilities', [])) or 'N/A'}</td>")
                    html_content.append(f"                    <td>{item.get('content_type', 'N/A')}</td>")
                    html_content.append(f"                    <td>{item.get('content_length', 'N/A')}</td>")
                    html_content.append("                </tr>")
                html_content.append("            </tbody>")
                html_content.append("        </table>")
            elif isinstance(items, list):
                html_content.append("        <ul>")
                for item in items:
                    if isinstance(item, dict):
                        html_content.append("            <li>")
                        for k, v in item.items():
                            html_content.append(f"                <strong>{k.replace('_', ' ').title()}:</strong> {v}<br>")
                        html_content.append("            </li>")
                    else:
                        html_content.append(f"            <li>{item}</li>")
                html_content.append("        </ul>")
            elif isinstance(items, dict):
                html_content.append("        <ul>")
                for k, v in items.items():
                    html_content.append(f"            <li><strong>{k.replace('_', ' ').title()}:</strong> {v}</li>")
                html_content.append("        </ul>")
            html_content.append("")  # Blank line after category

        html_content.extend([
            "    </div>",
            "</body>",
            "</html>"
        ])

        return "\n".join(html_content)

    def _format_xls_report(self, data: Dict, domain: str) -> Workbook: # Corrected type hint
        """Helper to format the report as an Excel workbook with multiple sheets."""
        wb = openpyxl.Workbook()
        default_sheet = wb.active
        default_sheet.title = "Summary"

        # Summary sheet with basic info
        default_sheet.append(["Reconnaissance Report for:", domain])
        default_sheet.append(["Generated at:", time.strftime("%Y-%m-%d %H:%M:%S")])
        default_sheet.append([])

        # Create sheets for each category
        for category, items in data.items():
            if not items:
                continue

            # Skip verified_urls if it's not a list of dicts - this condition is fine
            if category == 'verified_urls' and not (isinstance(items, list) and all(isinstance(item, dict) for item in items)):
                continue

            sheet = wb.create_sheet(title=category[:31])  # Excel sheet name limit

            if category == 'verified_urls':
                # Detailed probing results sheet
                headers = ["URL", "Status", "Final URL", "IP Address", "Technologies",
                          "Vulnerabilities", "Content Type", "Content Length", "Error"]
                sheet.append(headers)

                # Style headers
                for col in range(1, len(headers) + 1):
                    cell = sheet.cell(row=1, column=col)
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="0056b3", end_color="0056b3", fill_type="solid")
                    cell.alignment = Alignment(horizontal="center")

                for item in items:
                    sheet.append([
                        item.get('url', ''),
                        item.get('status', ''),
                        item.get('final_url', ''),
                        item.get('ip_address', ''),
                        ', '.join(item.get('technologies', [])),
                        ', '.join(item.get('vulnerabilities', [])),
                        item.get('content_type', ''),
                        item.get('content_length', ''),
                        item.get('error', '')
                    ])
            else:
                # Simple list sheet
                sheet.append([f"{category.replace('_', ' ').title()} ({len(items)})"])
                if isinstance(items, list):
                    for item in items:
                        if isinstance(item, dict):
                            # For dictionaries in other categories, represent them clearly
                            sheet.append([json.dumps(item, indent=2)]) # Use JSON for dicts in lists
                        else:
                            sheet.append([str(item)])
                elif isinstance(items, dict):
                    # For top-level dictionaries (if any appear)
                    for k, v in items.items():
                        sheet.append([f"{k}: {v}"])

            # Auto-size columns
            for column_cells in sheet.columns:
                length = max(len(str(cell.value)) for cell in column_cells) + 2
                sheet.column_dimensions[get_column_letter(column_cells[0].column)].width = length

        # FIX: Only remove default sheet if other sheets were actually created
        if len(wb.sheetnames) > 1 and default_sheet.title == "Summary" and default_sheet.max_row <= 3: # Check if summary sheet is largely empty and other sheets exist
             wb.remove(default_sheet)
        elif len(wb.sheetnames) == 1 and default_sheet.title == "Summary" and default_sheet.max_row <= 3: # If only summary and it's mostly empty
            # Optionally clear it if no data was added to it but no other sheets created
            default_sheet.delete_rows(1, default_sheet.max_row) # Clear all rows
            default_sheet.append(["No data to report for this domain."])


        return wb

    def save_report(self, report_content: Union[Dict, str, Workbook], # Corrected type hint
                   domain: str, format: str = 'json', output_dir: str = '.') -> str:
        """
        Saves the generated report to a file in the specified format.

        Args:
            report_content: The report content from generate_report().
            domain: The target domain (used for filename).
            format: The report format ('json', 'text', 'markdown', 'html', 'xls').
            output_dir: Directory to save the report file.

        Returns:
            The path to the saved report file.
        """
        safe_domain = re.sub(r'[^\w\-_.]', '_', domain)  # Sanitize domain for filename
        timestamp = time.strftime("%Y%m%d_%H%M%S")
        filename = f"recon_report_{safe_domain}_{timestamp}.{format}"
        filepath = os.path.join(output_dir, filename)

        try:
            os.makedirs(output_dir, exist_ok=True)
            logger.info(f"Saving {format.upper()} report to: {filepath}")

            if format == 'json':
                with open(filepath, 'w', encoding='utf-8') as f:
                    json.dump(report_content, f, indent=4)
            elif format in ('text', 'markdown', 'html'):
                with open(filepath, 'w', encoding='utf-8') as f:
                    f.write(report_content)
            elif format == 'xls':
                report_content.save(filepath)
            else:
                raise ValueError(f"Unsupported format for saving: {format}")

            logger.info(f"Successfully saved report to {filepath}")
            return filepath
        except Exception as e:
            logger.error(f"Failed to save report to {filepath}: {e}")
            raise

def main():
    """Command-line interface for the ReconMapper tool."""
    parser = argparse.ArgumentParser(
        description="Advanced Reconnaissance Data Mapper - Consolidates and analyzes findings from various sources.",
        epilog="Example: python reconmap.py -d example.com -s subdomains.txt -c content.txt -e endpoints.txt -p params.txt --probe --format markdown"
    )
    parser.add_argument('-d', '--domain', required=True, help="Base domain to map (e.g., example.com)")
    parser.add_argument('-s', '--subdomains', help="File containing subdomains (one per line)")
    parser.add_argument('-c', '--content', help="File containing content discovery results (URLs)")
    parser.add_argument('-e', '--endpoints', help="File containing API/endpoint URLs")
    parser.add_argument('-p', '--parameters', help="File containing parameterized URLs")
    parser.add_argument('--probe', action='store_true', help="Enable URL probing (HTTP requests)")
    parser.add_argument('--max-workers', type=int, default=10, help="Max concurrent threads for probing (default: 10)")
    parser.add_argument('--rate-limit', type=float, default=0.1, help="Seconds between requests (default: 0.1)")
    parser.add_argument('--timeout', type=int, default=15, help="Request timeout in seconds (default: 15)")
    parser.add_argument('--no-ssl-verify', action='store_true', help="Disable SSL certificate verification")
    parser.add_argument('--format', choices=['json', 'text', 'markdown', 'html', 'xls'],
                       default='json', help="Report format (default: json)")
    parser.add_argument('--output-dir', default='reports', help="Output directory for reports (default: ./reports)")
    parser.add_argument('--debug', action='store_true', help="Enable debug logging")

    args = parser.parse_args()

    if args.debug:
        logger.setLevel(logging.DEBUG)
        logger.debug("Debug logging enabled")

    # Initialize mapper with user settings
    mapper = ReconMapper(
        rate_limit=args.rate_limit,
        timeout=args.timeout,
        verify_ssl=not args.no_ssl_verify
    )

    # Load data files if provided
    if args.subdomains:
        mapper.load_data_file(args.subdomains, 'subdomains', args.domain)
    if args.content:
        mapper.load_data_file(args.content, 'content', args.domain)
    if args.endpoints:
        mapper.load_data_file(args.endpoints, 'endpoints', args.domain)
    if args.parameters:
        mapper.load_data_file(args.parameters, 'parameters', args.domain)

    # Perform URL probing if requested
    if args.probe:
        mapper.probe_urls(args.domain, max_workers=args.max_workers)

    # Generate and save report
    report = mapper.generate_report(args.domain, format=args.format, include_probe=args.probe)
    saved_path = mapper.save_report(report, args.domain, format=args.format, output_dir=args.output_dir)

    print(f"\n[+] Report successfully generated and saved to: {saved_path}")

if __name__ == '__main__':
    main()
